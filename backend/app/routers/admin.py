from fastapi import (
    APIRouter,
    Depends,
    UploadFile,
    File,
    HTTPException,
    Form,
    Query,
)
import io
from sqlalchemy.orm import Session
from sqlalchemy import text,select,func
from sqlalchemy.exc import IntegrityError
import csv
from openpyxl import load_workbook
import os
from io import StringIO
import logging
from typing import List, Dict
from fastapi.responses import StreamingResponse
import json
import re

from app.deps import get_db
from app import models
from app import schemas
from app.models import (
    CareerCluster,
    Career,
    KeySkill,
    Student,
    Skill,
    StudentSkillMap,
    StudentKeySkillMap,
    User as UserModel,
    career_keyskill_association,
    Question,
    ExplainabilityContent,
    
)
from app.schemas import (
    UploadResponse,
    UploadQuestionsResult,
    RoleChange,
    GuardianAssign,
    User as UserSchema,
    AdminQuestionCreateRequest,
    AdminQuestionCreateResponse,
    AdminQuestionBulkItem,
    AdminQuestionBulkResponse,
    AdminQuestionBulkErrorEntry,
    ValidateKnowledgePackResponse,
    ExplainabilityUploadResult,
    ExplainabilityUploadRowError,
    ValidateExplainabilityKeysResponse,
    ExplainabilityCoverageResponse,
)
from app.auth.auth import require_role, get_current_active_user
from app.services.knowledge_pack_validation import (
    run_validate_knowledge_pack,
    run_validate_explainability_keys,
)

from app.services.skill_keyskill_ingest import ingest_skill_keyskill_map
from app.utils.alias_normalization import resolve_alias
from app.utils.normalization import norm

# ✅ B2 shared validation import
from app.validators.question_ingestion import (
    validate_question_row,
    RowValidationError,
)

logger = logging.getLogger(__name__)

router = APIRouter(
    tags=["admin"],
    # Enforce admin at router level
    dependencies=[Depends(require_role("admin"))],
)

@router.post(
    "/upload-skill-keyskill-map",
    summary="PR46: Upload StudentSkill → KeySkill semantic map from Excel (dry_run supported)",
)
async def upload_skill_keyskill_map(
    file: UploadFile = File(...),
    dry_run: bool = Query(True),
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    # Basic file validation
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx Excel files are accepted")

    # Read file bytes ONCE
    content = await file.read()

    # Guard: empty bytes usually means stream was consumed or file invalid
    if not content or len(content) < 200:
        raise HTTPException(
            status_code=400,
            detail="Uploaded file content is empty/invalid. Please re-select the .xlsx file and try again.",
        )

    try:
        return ingest_skill_keyskill_map(db=db, file_bytes=content, dry_run=dry_run)
    except ValueError as ve:
        raise HTTPException(status_code=422, detail=str(ve))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

def _parse_optional_int(value):
    """
    Convert optional values to int for DB integer columns.

    Accepts:
      - None, "", "null" -> None
      - "12" -> 12
      - 12 -> 12

    Raises:
      - ValueError for non-numeric strings.
    """
    if value is None:
        return None
    if isinstance(value, int):
        return value

    s = str(value).strip()
    if s == "" or s.lower() == "null":
        return None

    return int(s)


# ============================================================
# 1. UPLOAD CAREER CLUSTERS
# ============================================================

@router.post(
    "/upload-career-clusters",
    response_model=UploadResponse,
    summary="Bulk upload Career Clusters via CSV",
)
async def upload_career_clusters(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    """
    Expected CSV headers (exact):
      cluster_id,cluster_name

    Behavior:
    - If cluster exists, update name.
    - Else insert.
    """
    if file.content_type != "text/csv":
        raise HTTPException(status_code=400, detail="Only text/csv files are accepted")

    text = (await file.read()).decode("utf-8-sig")
    if not text.strip():
        raise HTTPException(status_code=400, detail="CSV file is empty")

    reader = csv.DictReader(io.StringIO(text))

    expected_fields = {"cluster_id", "cluster_name"}
    if set(reader.fieldnames or []) != expected_fields:
        raise HTTPException(
            status_code=400,
            detail=f"CSV must have columns exactly {expected_fields}, but got {reader.fieldnames}",
        )

    inserted = 0
    updated = 0
    skipped = 0

    for row in reader:
        cid = (row.get("cluster_id") or "").strip()
        cname = (row.get("cluster_name") or "").strip()

        if not cid or not cname:
            skipped += 1
            continue

        existing = db.query(CareerCluster).filter_by(cluster_id=cid).first()
        if existing:
            if (existing.cluster_name or "").strip() != cname:
                existing.cluster_name = cname
                updated += 1
            else:
                skipped += 1
            continue

        db.add(CareerCluster(cluster_id=cid, cluster_name=cname))
        inserted += 1

    db.commit()
    logger.info(f"upload-career-clusters: inserted={inserted}, updated={updated}, skipped={skipped}")
    return {"status": "success", "inserted": inserted}


# ============================================================
# 2. UPLOAD CAREERS
# ============================================================

@router.post(
    "/upload-careers",
    response_model=UploadResponse,
    summary="Bulk upload Careers via CSV",
)
async def upload_careers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    """
    Expected CSV headers (exact):
      career_id,career_name,cluster_id

    Behavior:
    - If career exists, update fields.
    - Else insert.
    - cluster_id must exist.
    """
    if file.content_type != "text/csv":
        raise HTTPException(status_code=400, detail="Only text/csv files are accepted")

    text = (await file.read()).decode("utf-8-sig")
    if not text.strip():
        raise HTTPException(status_code=400, detail="CSV file is empty")

    reader = csv.DictReader(io.StringIO(text))

    expected_fields = {"career_id", "career_name", "cluster_id"}
    if set(reader.fieldnames or []) != expected_fields:
        raise HTTPException(
            status_code=400,
            detail=f"CSV must have columns exactly {expected_fields}, but got {reader.fieldnames}",
        )

    inserted = 0
    updated = 0
    skipped = 0

    for row in reader:
        career_id = (row.get("career_id") or "").strip()
        career_name = (row.get("career_name") or "").strip()
        cluster_id = (row.get("cluster_id") or "").strip()

        if not career_id or not career_name or not cluster_id:
            skipped += 1
            continue

        cluster = db.query(CareerCluster).filter_by(cluster_id=cluster_id).first()
        if not cluster:
            skipped += 1
            continue

        existing = db.query(Career).filter_by(career_id=career_id).first()
        if existing:
            changed = False
            if (existing.career_name or "").strip() != career_name:
                existing.career_name = career_name
                changed = True
            if (existing.cluster_id or "").strip() != cluster_id:
                existing.cluster_id = cluster_id
                changed = True

            if changed:
                updated += 1
            else:
                skipped += 1
            continue

        db.add(Career(career_id=career_id, career_name=career_name, cluster_id=cluster_id))
        inserted += 1

    db.commit()
    logger.info(f"upload-careers: inserted={inserted}, updated={updated}, skipped={skipped}")
    return {"status": "success", "inserted": inserted}


# ============================================================
# 3. UPLOAD KEY SKILLS
# ============================================================

@router.post(
    "/upload-keyskills",
    response_model=UploadResponse,
    summary="Bulk upload KeySkills via CSV",
)
async def upload_keyskills(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    """
    Expected CSV headers (exact):
      keyskill_id,keyskill_name

    Behavior:
    - If exists, update name.
    - Else insert.
    """
    if file.content_type != "text/csv":
        raise HTTPException(status_code=400, detail="Only text/csv files are accepted")

    text = (await file.read()).decode("utf-8-sig")
    if not text.strip():
        raise HTTPException(status_code=400, detail="CSV file is empty")

    reader = csv.DictReader(io.StringIO(text))

    expected_fields = {"keyskill_id", "keyskill_name"}
    if set(reader.fieldnames or []) != expected_fields:
        raise HTTPException(
            status_code=400,
            detail=f"CSV must have columns exactly {expected_fields}, but got {reader.fieldnames}",
        )

    inserted = 0
    updated = 0
    skipped = 0

    for row in reader:
        kid = (row.get("keyskill_id") or "").strip()
        kname = (row.get("keyskill_name") or "").strip()

        if not kid or not kname:
            skipped += 1
            continue

        existing = db.query(KeySkill).filter_by(keyskill_id=kid).first()
        if existing:
            if (existing.keyskill_name or "").strip() != kname:
                existing.keyskill_name = kname
                updated += 1
            else:
                skipped += 1
            continue

        db.add(KeySkill(keyskill_id=kid, keyskill_name=kname))
        inserted += 1

    db.commit()
    logger.info(f"upload-keyskills: inserted={inserted}, updated={updated}, skipped={skipped}")
    return {"status": "success", "inserted": inserted}


# ============================================================
# 4. CAREER ↔ KEYSKILL MAPPING
# ============================================================

@router.post(
    "/upload-career-keyskill-map",
    response_model=UploadResponse,
    summary="Bulk upload Career ↔ KeySkill mapping via CSV",
)
async def upload_career_keyskill_map(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    """
    Expected CSV headers (exact):
      career_id,keyskill_id

    Behavior:
    - Inserts into association table.
    - Skips invalid FK rows.
    - Skips duplicates (idempotent-ish).
    """
    if file.content_type != "text/csv":
        raise HTTPException(status_code=400, detail="Only text/csv files are accepted")

    text = (await file.read()).decode("utf-8-sig")
    if not text.strip():
        raise HTTPException(status_code=400, detail="CSV file is empty")

    reader = csv.DictReader(io.StringIO(text))

    expected_fields = {"career_id", "keyskill_id"}
    if set(reader.fieldnames or []) != expected_fields:
        raise HTTPException(
            status_code=400,
            detail=f"CSV must have columns exactly {expected_fields}, but got {reader.fieldnames}",
        )

    inserted = 0
    skipped = 0

    for row in reader:
        career_id = (row.get("career_id") or "").strip()
        keyskill_id = (row.get("keyskill_id") or "").strip()

        if not career_id or not keyskill_id:
            skipped += 1
            continue

        career = db.query(Career).filter_by(career_id=career_id).first()
        keyskill = db.query(KeySkill).filter_by(keyskill_id=keyskill_id).first()
        if not career or not keyskill:
            skipped += 1
            continue

        # Insert into association table if not exists
        exists = db.execute(
            career_keyskill_association.select().where(
                (career_keyskill_association.c.career_id == career_id)
                & (career_keyskill_association.c.keyskill_id == keyskill_id)
            )
        ).first()

        if exists:
            skipped += 1
            continue

        db.execute(
            career_keyskill_association.insert().values(
                career_id=career_id, keyskill_id=keyskill_id
            )
        )
        inserted += 1

    db.commit()
    logger.info(f"upload-career-keyskill-map: inserted={inserted}, skipped={skipped}")
    return {"status": "success", "inserted": inserted}

# ============================================================
# PR36. UPLOAD CAREER ↔ KEYSKILL WEIGHTS (STRICT SUM=100)
# ============================================================

@router.post(
    "/upload-career-keyskill-weights",
    summary="Bulk upload Career ↔ KeySkill weights via CSV (sum=100 per career, strict)",
)
async def upload_career_keyskill_weights(
    file: UploadFile = File(...),
    dry_run: bool = False,
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    """
    PR36:
    - Admin-only (router dependency already enforces)
    - CSV headers (exact): career_id,keyskill_id,weight_percentage
    - Validates:
        * ints for ids + weight
        * weight >= 0
        * all (career_id) exist in careers.id
        * all (keyskill_id) exist in keyskills.id
        * per career_id: SUM(weight_percentage) == 100
    - Behavior:
        * If ANY career fails sum=100 => 400 with list of bad careers
        * If ok => idempotent upsert into career_keyskill_association
        * dry_run=true => validate + count inserts/updates, but DO NOT write
    """

    # --- file type checks ---
    if not (file.filename or "").lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only .csv files are accepted")

    raw = await file.read()

    # Robust decoding: prefer UTF-8 BOM-safe, fallback to UTF-16 (Excel common)
    try:
        text_csv = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text_csv = raw.decode("utf-16")
        except UnicodeDecodeError:
            raise HTTPException(
                status_code=400,
                detail="Unable to decode CSV. Please save as CSV UTF-8 (Comma delimited).",
            )

    if not text_csv.strip():
        raise HTTPException(status_code=400, detail="CSV file is empty")

    reader = csv.DictReader(io.StringIO(text_csv))

    expected_headers = ["career_id", "keyskill_id", "weight_percentage"]
    actual_headers = reader.fieldnames or []
    if actual_headers != expected_headers:
        raise HTTPException(
            status_code=400,
            detail=f"Header mismatch. Expected exactly {expected_headers} but got {actual_headers}",
        )

    rows = []
    errors = []

    # ---------------------------
    # PASS 1: Parse + row-level validation (no DB writes)
    # ---------------------------
    for line_no, r in enumerate(reader, start=2):
        try:
            career_id_raw = (r.get("career_id") or "").strip()
            keyskill_id_raw = (r.get("keyskill_id") or "").strip()
            weight_raw = (r.get("weight_percentage") or "").strip()

            if not career_id_raw or not keyskill_id_raw or weight_raw == "":
                raise ValueError("career_id, keyskill_id, weight_percentage are required")

            career_id = int(career_id_raw)
            keyskill_id = int(keyskill_id_raw)
            weight = int(weight_raw)

            if weight < 0:
                raise ValueError("weight_percentage must be >= 0")

            rows.append(
                {
                    "rownum": line_no,
                    "career_id": career_id,
                    "keyskill_id": keyskill_id,
                    "weight_percentage": weight,
                }
            )

        except Exception as e:
            errors.append({"row": line_no, "error": str(e), "raw": dict(r)})

    if errors:
        raise HTTPException(status_code=400, detail={"ok": False, "errors": errors})

    # ---------------------------
    # PASS 2: Duplicate detection inside file (career_id, keyskill_id)
    # ---------------------------
    seen = set()
    dupes = []
    for r in rows:
        k = (r["career_id"], r["keyskill_id"])
        if k in seen:
            dupes.append({"row": r["rownum"], "error": "Duplicate (career_id, keyskill_id) in file"})
        seen.add(k)

    if dupes:
        raise HTTPException(status_code=400, detail={"ok": False, "errors": dupes})

    # ---------------------------
    # PASS 3: FK existence checks (careers + keyskills)
    # ---------------------------
    career_ids = sorted({r["career_id"] for r in rows})
    keyskill_ids = sorted({r["keyskill_id"] for r in rows})

    career_ok_rows = db.execute(
        text("SELECT id FROM careers WHERE id = ANY(:ids)"),
        {"ids": career_ids},
    ).fetchall()
    career_ok = {x[0] for x in career_ok_rows}

    keyskill_ok_rows = db.execute(
        text("SELECT id FROM keyskills WHERE id = ANY(:ids)"),
        {"ids": keyskill_ids},
    ).fetchall()
    keyskill_ok = {x[0] for x in keyskill_ok_rows}

    fk_errors = []
    for r in rows:
        if r["career_id"] not in career_ok:
            fk_errors.append({"row": r["rownum"], "field": "career_id", "error": f"career_id not found: {r['career_id']}"})
        if r["keyskill_id"] not in keyskill_ok:
            fk_errors.append({"row": r["rownum"], "field": "keyskill_id", "error": f"keyskill_id not found: {r['keyskill_id']}"})

    if fk_errors:
        raise HTTPException(status_code=400, detail={"ok": False, "errors": fk_errors})

    # ---------------------------
    # PASS 4: Group-level validation (sum=100 per career)
    # ---------------------------
    sums = {}
    for r in rows:
        sums[r["career_id"]] = sums.get(r["career_id"], 0) + r["weight_percentage"]

    bad = [{"career_id": cid, "sum_weight": s} for cid, s in sums.items() if s != 100]
    if bad:
        # strict policy: reject the entire upload if any career fails
        raise HTTPException(
            status_code=400,
            detail={
                "ok": False,
                "error_code": "CAREER_KEYSKILL_WEIGHT_SUM_NOT_100",
                "message": "One or more careers have weight_percentage sum != 100",
                "bad_careers": sorted(bad, key=lambda x: x["career_id"]),
            },
        )

    # ---------------------------
    # PASS 5: Upsert (idempotent)
    # ---------------------------
    inserted = 0
    updated = 0

    for r in rows:
        # check existence for accurate inserted vs updated counts
        exists = db.execute(
            text(
                """
                SELECT 1
                FROM career_keyskill_association
                WHERE career_id = :c AND keyskill_id = :k
                LIMIT 1
                """
            ),
            {"c": r["career_id"], "k": r["keyskill_id"]},
        ).first()

        if dry_run:
            if exists:
                updated += 1
            else:
                inserted += 1
            continue

        db.execute(
            text(
                """
                INSERT INTO career_keyskill_association (career_id, keyskill_id, weight_percentage)
                VALUES (:career_id, :keyskill_id, :weight_percentage)
                ON CONFLICT (career_id, keyskill_id)
                DO UPDATE SET weight_percentage = EXCLUDED.weight_percentage
                """
            ),
            r,
        )

        if exists:
            updated += 1
        else:
            inserted += 1

    if not dry_run:
        db.commit()

    return {
        "ok": True,
        "dry_run": dry_run,
        "inserted": inserted,
        "updated": updated,
        "total_rows": len(rows),
    }
# ============================================================
# ✅ B3. API-BASED QUESTION CREATION (JSON)
# ============================================================

@router.post(
    "/questions",
    response_model=AdminQuestionCreateResponse,
    status_code=201,
    summary="Create a single Question via API (JSON) using shared validator (B3)",
)
def create_question_api(
    payload: AdminQuestionCreateRequest,
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    """
    B3 spec:
    - Admin-only (router dependency)
    - Input: JSON body
    - Reuses shared validate_question_row() (B2)
    - No DB writes before validation passes
    - Duplicate question_id => 409 Conflict (explicit error)

    Important:
    - `id` is the internal DB PK (integer)
    - `question_code` is the canonical/external identifier used by student flows
    """

    # Convert payload -> dict to match validator signature
    row_for_validation = payload.dict()

    validated = validate_question_row(db=db, row=row_for_validation, row_index=1)

    # Validation failure (400) - do not write anything to DB
    if isinstance(validated, RowValidationError):
        raise HTTPException(
            status_code=400,
            detail={
                "status": "error",
                "error_code": "ROW_VALIDATION_ERROR",
                "errors": [{"field": e.field, "message": e.message} for e in validated.errors],
            },
        )

    # DB write ONLY after validation succeeds
    try:
        qid_int = int(str(payload.question_id).strip())
    except Exception:
        raise HTTPException(
            status_code=400,
            detail={
                "status": "error",
                "error_code": "INVALID_QUESTION_ID",
                "message": f"question_id must be an integer-like value (got '{payload.question_id}')",
                "field": "question_id",
            },
        )

    # prerequisite_qid is stored as INTEGER FK -> must be int/None
    try:
        prereq_int = _parse_optional_int(payload.prerequisite_qid)
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail={
                "status": "error",
                "error_code": "INVALID_PREREQUISITE_QID",
                "message": "prerequisite_qid must be an integer (or empty/null)",
                "field": "prerequisite_qid",
            },
        )

    q = models.Question(
        id=qid_int,
        assessment_version=validated.assessment_version,
        question_text_en=validated.question_text_en,
        question_text_hi=(validated.question_text_hi or "").strip() or None,
        question_text_ta=(validated.question_text_ta or "").strip() or None,
        skill_id=validated.skill_id,

        # Optional fields supported by your questions table
        weight=payload.weight if payload.weight is not None else 1,
        group_id=(payload.group_id or "").strip() or None,
        prerequisite_qid=prereq_int,

        # Canonical external identifier (must be persisted if supplied)
        question_code=(payload.question_code or "").strip() or None,
    )

    db.add(q)
    try:
        db.commit()
        db.refresh(q)
    except IntegrityError:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail={
                "status": "error",
                "error_code": "DUPLICATE_QUESTION_ID",
                "message": f"Question with id '{qid_int}' already exists.",
                "field": "question_id",
            },
        )
    except Exception as e:
        db.rollback()
        logger.exception("create_question_api: failed DB write")
        raise HTTPException(status_code=500, detail=f"Failed to create question: {e}")

    return AdminQuestionCreateResponse(
        status="created",
        created={
            "question_id": q.id,
            "assessment_version": q.assessment_version,
            "skill_id": q.skill_id,
        },
        errors=[],
    )


# ============================================================
# ✅ B4. API-BASED BULK QUESTION CREATION (JSON ARRAY)
# ============================================================

@router.post(
    "/questions/bulk",
    response_model=AdminQuestionBulkResponse,
    status_code=200,
    summary="Create Questions in bulk via API (JSON array) using shared validator (B4)",
)
def bulk_create_questions_api(
    payloads: List[AdminQuestionBulkItem],
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    """
    B4 spec:
    - Admin-only (router dependency)
    - Input: JSON ARRAY
    - Validates first (no writes), writes later
    - Continues on error
    - Skips duplicates (both within request + existing DB rows)
    - Must persist `question_code` (canonical external identifier)
    - Must store `prerequisite_qid` as int/None (INTEGER FK column)
    """

    created = 0
    skipped = 0
    errors: List[AdminQuestionBulkErrorEntry] = []

    # ---------------------------
    # Pre-check payload-level duplicates for question_id (string compare after strip)
    # ---------------------------
    seen_qids = set()
    duplicate_indexes = set()

    for i, item in enumerate(payloads):
        qid = (item.question_id or "").strip()
        if not qid:
            duplicate_indexes.add(i)
            continue
        if qid in seen_qids:
            duplicate_indexes.add(i)
        else:
            seen_qids.add(qid)

    # ---------------------------
    # PASS 1: Validate & normalize (NO DB WRITES)
    # ---------------------------
    valid_items: List[Dict] = []   # insert-ready dicts
    valid_meta: List[Dict] = []    # index + question_id for error reporting

    for i, item in enumerate(payloads):
        qid = (item.question_id or "").strip()

        # Skip payload-level duplicates
        if i in duplicate_indexes:
            skipped += 1
            errors.append(
                AdminQuestionBulkErrorEntry(
                    index=i,
                    question_id=qid or None,
                    errors=[{"field": "question_id", "message": "duplicate question_id in request payload"}],
                )
            )
            continue

        # Convert to dict to match validator signature
        row_for_validation = item.dict()
        validated = validate_question_row(db=db, row=row_for_validation, row_index=i)

        if isinstance(validated, RowValidationError):
            skipped += 1
            errors.append(
                AdminQuestionBulkErrorEntry(
                    index=i,
                    question_id=qid or None,
                    errors=[{"field": e.field, "message": e.message} for e in validated.errors],
                )
            )
            continue

        # id is INTEGER PK in DB
        try:
            qid_int = int(qid)
        except Exception:
            skipped += 1
            errors.append(
                AdminQuestionBulkErrorEntry(
                    index=i,
                    question_id=qid or None,
                    errors=[{"field": "question_id", "message": f"must be an integer-like value (got '{qid}')"}],
                )
            )
            continue

        # prerequisite_qid is INTEGER FK -> must be int/None
        try:
            prereq_int = _parse_optional_int(item.prerequisite_qid)
        except ValueError:
            skipped += 1
            errors.append(
                AdminQuestionBulkErrorEntry(
                    index=i,
                    question_id=qid or None,
                    errors=[{"field": "prerequisite_qid", "message": "must be an integer (or empty/null)"}],
                )
            )
            continue

        # Build insert-ready dict (still NO DB write)
        valid_items.append(
            {
                "id": qid_int,
                "assessment_version": validated.assessment_version,
                "question_text_en": validated.question_text_en,
                "question_text_hi": (validated.question_text_hi or "").strip() or None,
                "question_text_ta": (validated.question_text_ta or "").strip() or None,
                "skill_id": validated.skill_id,
                "weight": item.weight if item.weight is not None else 1,
                "group_id": (item.group_id or "").strip() or None,
                "prerequisite_qid": prereq_int,
                # ✅ Persist canonical identifier (external)
                "question_code": (getattr(item, "question_code", None) or "").strip() or None,
            }
        )
        valid_meta.append({"index": i, "question_id": qid})

    # ---------------------------
    # PASS 2: Write valid rows (DB writes happen ONLY here)
    # Flush each row to catch per-row IntegrityError without killing whole batch.
    # ---------------------------
    for meta, row in zip(valid_meta, valid_items):
        qid = meta["question_id"]

        q = models.Question(**row)
        db.add(q)

        try:
            db.flush()  # detect duplicates early
            created += 1

        except IntegrityError:
            db.rollback()
            skipped += 1
            errors.append(
                AdminQuestionBulkErrorEntry(
                    index=meta["index"],
                    question_id=qid,
                    errors=[{"field": "question_id", "message": "already exists"}],
                )
            )

        except Exception as e:
            db.rollback()
            skipped += 1
            logger.exception("bulk_create_questions_api: unexpected DB error")
            errors.append(
                AdminQuestionBulkErrorEntry(
                    index=meta["index"],
                    question_id=qid,
                    errors=[{"field": "db", "message": f"failed to insert: {e}"}],
                )
            )

    # Final commit for the successful inserts
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        logger.exception("bulk_create_questions_api: commit failed")
        raise HTTPException(status_code=500, detail=f"Bulk insert commit failed: {e}")

    return AdminQuestionBulkResponse(created=created, skipped=skipped, errors=errors)


# ============================================================
# 5. STUDENT SKILL MAPS
# ============================================================

@router.post(
    "/student-skill-map",
    summary="Bulk map Student ↔ Skill",
)
def student_skill_map(
    mappings: List[Dict[str, int]],
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    inserted = 0
    for item in mappings:
        sid = item.get("student_id")
        kid = item.get("skill_id")
        if not db.query(Student).get(sid) or not db.query(Skill).get(kid):
            raise HTTPException(status_code=400, detail="Invalid student_id or skill_id")

        db.add(StudentSkillMap(student_id=sid, skill_id=kid))
        inserted += 1

    db.commit()
    return {"status": "success", "inserted": inserted}


@router.post(
    "/student-keyskill-map",
    summary="Bulk map Student ↔ KeySkill",
)
def student_keyskill_map(
    mappings: List[Dict[str, int]],
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    inserted = 0
    for item in mappings:
        sid = item.get("student_id")
        kkid = item.get("keyskill_id")
        if not db.query(Student).get(sid) or not db.query(KeySkill).get(kkid):
            raise HTTPException(status_code=400, detail="Invalid student_id or keyskill_id")

        db.add(StudentKeySkillMap(student_id=sid, keyskill_id=kkid))
        inserted += 1

    db.commit()
    return {"status": "success", "inserted": inserted}


# ============================================================
# 6. USER MANAGEMENT
# ============================================================

@router.get(
    "/list-users",
    response_model=List[UserSchema],
    summary="List all users",
)
def list_users(
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    return db.query(UserModel).all()


@router.post(
    "/change-role/{user_id}",
    response_model=UserSchema,
    summary="Change a user’s role",
)
def change_role(
    user_id: int,
    payload: RoleChange,
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    user = db.query(UserModel).get(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    user.role = payload.role
    db.commit()
    db.refresh(user)

    return user


@router.post(
    "/assign-guardian/{user_id}",
    response_model=UserSchema,
    summary="Assign guardian to a user",
)
def assign_guardian(
    user_id: int,
    payload: GuardianAssign,
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    user = db.query(UserModel).get(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    user.guardian_email = payload.guardian_email
    db.commit()
    db.refresh(user)

    return user

@router.get(
    "/validate-knowledge-pack",
    tags=["Admin Panel", "admin"],
)
def validate_knowledge_pack(
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    # Import-safe: keep schemas inside the service to avoid startup failures
    return run_validate_knowledge_pack(db)

@router.get(
    "/validate-knowledge-pack.csv",
    tags=["Admin Panel", "admin"],
)
def validate_knowledge_pack_csv(
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    """
    PR12: Admin productivity
    - Export the validation issues as CSV for remediation workflows.
    - Read-only; does not modify ingestion/scoring.
    """
    report = run_validate_knowledge_pack(db)

    # CSV headers (stable + remediation-friendly)
    headers = ["severity", "code", "message", "sample_json"]

    def _iter_rows():
        # header row
        yield ",".join(headers) + "\n"
        for issue in report.issues:
            sev = (issue.severity or "").replace('"', '""')
            code = (issue.code or "").replace('"', '""')
            msg = (issue.message or "").replace('"', '""')

            # sample can be dict/None; store as compact JSON string
            sample_obj = issue.sample if issue.sample is not None else None
            sample_json = json.dumps(sample_obj, ensure_ascii=False, separators=(",", ":"))
            sample_json = sample_json.replace('"', '""')  # CSV escape

            yield f'"{sev}","{code}","{msg}","{sample_json}"\n'

    return StreamingResponse(
        _iter_rows(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=knowledge_pack_validation_issues.csv"},
    )

@router.get(
    "/validate-explainability-keys",
    response_model=ValidateExplainabilityKeysResponse,
    tags=["Admin Panel", "admin"],
    summary="PR39: Validate explainability_key taxonomy + coverage (read-only)",
)
def validate_explainability_keys(
    version: str | None = Query(default=None),
    locale: str | None = Query(default=None),
    required_families: str | None = Query(
        default=None,
        description="Comma-separated family list (default: AQ,FACET,SKILL,CAREER,CLUSTER)",
    ),
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    families = None
    if required_families:
        families = [x.strip().upper() for x in required_families.split(",") if x.strip()]

    return run_validate_explainability_keys(
        db=db,
        version=version,
        locale=locale,
        required_families=families,
    )
# ============================================================
# PR41 — Locale Coverage Validation (i18n Gate)
# ============================================================
@router.get(
    "/explainability-coverage",
    response_model=schemas.ExplainabilityCoverageResponse,
    tags=["Admin Panel", "admin"],
    summary="PR41: Coverage report for missing explanation keys per locale/version (baseline en)",
)
def explainability_coverage(
    version: str = Query(..., min_length=1, max_length=32),
    locale: str = Query(..., min_length=1, max_length=20),
    baseline_locale: str = Query("en", min_length=1, max_length=20),
    format: str = Query("json", description="json | csv"),
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    if getattr(current_user, "role", None) != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    # ---- Pull active keys for baseline + target ----
    baseline_rows = (
        db.execute(
            select(ExplainabilityContent.explanation_key)
            .where(
                ExplainabilityContent.version == version,
                ExplainabilityContent.locale == baseline_locale,
                ExplainabilityContent.is_active == True,  # noqa: E712
            )
        )
        .scalars()
        .all()
    )

    target_rows = (
        db.execute(
            select(ExplainabilityContent.explanation_key)
            .where(
                ExplainabilityContent.version == version,
                ExplainabilityContent.locale == locale,
                ExplainabilityContent.is_active == True,  # noqa: E712
            )
        )
        .scalars()
        .all()
    )

    baseline_keys = set([k.strip() for k in baseline_rows if k])
    target_keys = set([k.strip() for k in target_rows if k])

    missing = sorted(list(baseline_keys - target_keys))

    payload = schemas.ExplainabilityCoverageResponse(
        version=version,
        locale=locale,
        baseline_locale=baseline_locale,
        baseline_active_keys=len(baseline_keys),
        target_active_keys=len(target_keys),
        missing_count=len(missing),
        missing_keys=missing,
    )

    # ---- CSV projection (exportable) ----
    if (format or "").lower() == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["version", "baseline_locale", "target_locale", "missing_explanation_key"])
        for k in missing:
            w.writerow([version, baseline_locale, locale, k])

        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="explainability_coverage_{version}_{locale}.csv"'
            },
        )

    return payload
# ============================================================
# PR45. QSSW UPLOAD
# ============================================================
@router.post(
    "/upload-question-student-skill-weights",
    response_model=schemas.QSSWUploadResult,
    summary="Upload Question → StudentSkill weights (QSSW) (idempotent upsert, supports dry_run)",
)
async def upload_question_student_skill_weights(
    file: UploadFile = File(...),
    dry_run: bool = False,
    default_assessment_version: str = "v1",
    db: Session = Depends(get_db),
):
    """
    CSV headers (minimum):
      canonical_student_skill, weight

    Plus ONE of:
      question_id   (numeric questions.id)
      question_code (stable external identifier)

    Optional:
      assessment_version (defaults to default_assessment_version)
      source, facet_id, aq_id

    Behavior:
    - Resolve question_id:
        if question_id provided: validate exists in questions.id
        else: resolve by (assessment_version, question_code)
    - Resolve student_skill via skills.name exact match (trimmed)
    - Upsert into question_student_skill_weights on (question_id, skill_id)
    - dry_run=true: validate + count what would happen, but DO NOT write
    - Returns inserted/updated/skipped + per-row error list
    """

    # --- read + decode (Excel safe) ---
    raw = await file.read()
    try:
        text_csv = raw.decode("utf-8-sig")   # handles UTF-8 BOM from Excel
    except UnicodeDecodeError:
        text_csv = raw.decode("utf-16")      # common Excel fallback

    reader = csv.DictReader(StringIO(text_csv))
    if not reader.fieldnames:
        return schemas.QSSWUploadResult(ok=False, dry_run=dry_run, errors=[
            schemas.QSSWUploadRowError(row=1, error="CSV missing header row", raw={})
        ])

    headers = {h.strip() for h in reader.fieldnames if h}

    required = {"canonical_student_skill", "weight"}
    missing = required - headers
    if missing:
        return schemas.QSSWUploadResult(ok=False, dry_run=dry_run, errors=[
            schemas.QSSWUploadRowError(row=1, error=f"Missing required columns: {sorted(missing)}", raw={})
        ])

    if "question_id" not in headers and "question_code" not in headers:
        return schemas.QSSWUploadResult(ok=False, dry_run=dry_run, errors=[
            schemas.QSSWUploadRowError(row=1, error="CSV must include either question_id or question_code", raw={})
        ])

    # --- pre-load skill name -> id map (fast + deterministic) ---
    skill_rows = db.query(models.Skill.id, models.Skill.name).all()
    skill_by_name = {(name or "").strip(): sid for sid, name in skill_rows}

    inserted = 0
    updated = 0
    skipped = 0
    errors = []

    # cache question lookups by (assessment_version, question_code)
    q_cache = {}

    # NOTE: csv line numbers start at 2 because header is row 1
    for line_no, r in enumerate(reader, start=2):
        try:
            av = (r.get("assessment_version") or "").strip() or default_assessment_version

            qid_raw = (r.get("question_id") or "").strip()
            qcode = (r.get("question_code") or "").strip()

            # --- resolve question_id ---
            question_id = None

            if qid_raw:
                if not qid_raw.isdigit():
                    raise ValueError("question_id must be numeric when provided")
                question_id = int(qid_raw)

                # validate question exists
                q_exists = db.query(models.Question.id).filter(models.Question.id == question_id).first()
                if not q_exists:
                    raise ValueError(f"question_id not found: {question_id}")

            else:
                if not qcode:
                    raise ValueError("Either question_id or question_code must be provided")

                key = (av, qcode)
                if key in q_cache:
                    question_id = q_cache[key]
                else:
                    q = (
                        db.query(models.Question.id)
                        .filter(models.Question.assessment_version == av, models.Question.question_code == qcode)
                        .first()
                    )
                    if not q:
                        raise ValueError(f"question_code not found for assessment_version={av}: {qcode}")
                    question_id = q[0]
                    q_cache[key] = question_id

            # --- resolve skill_id ---
            skill_name = (r.get("canonical_student_skill") or "").strip()
            if not skill_name:
                raise ValueError("canonical_student_skill is blank")

            skill_id = skill_by_name.get(skill_name)
            if not skill_id:
                raise ValueError(f"skill not found (skills.name): {skill_name}")

            # --- weight ---
            w_raw = (r.get("weight") or "").strip()
            if w_raw == "":
                raise ValueError("weight is blank")
            try:
                weight = float(w_raw)
            except Exception:
                raise ValueError(f"weight not numeric: {w_raw}")

            # optional metadata
            source = (r.get("source") or "").strip() or None
            facet_id = (r.get("facet_id") or "").strip() or None
            aq_id = (r.get("aq_id") or "").strip() or None

            # --- determine insert vs update (for counts) ---
            exists_row = db.query(models.QuestionStudentSkillWeight.id).filter(
                models.QuestionStudentSkillWeight.question_id == question_id,
                models.QuestionStudentSkillWeight.skill_id == skill_id,
            ).first()

            if dry_run:
                if exists_row:
                    updated += 1
                else:
                    inserted += 1
                continue

            # --- write: use ON CONFLICT for true idempotent upsert ---
            stmt = text("""
                INSERT INTO question_student_skill_weights
                    (question_id, skill_id, weight, source, facet_id, aq_id)
                VALUES
                    (:question_id, :skill_id, :weight, :source, :facet_id, :aq_id)
                ON CONFLICT (question_id, skill_id)
                DO UPDATE SET
                    weight = EXCLUDED.weight,
                    source = EXCLUDED.source,
                    facet_id = EXCLUDED.facet_id,
                    aq_id = EXCLUDED.aq_id
            """)

            db.execute(stmt, {
                "question_id": question_id,
                "skill_id": skill_id,
                "weight": weight,
                "source": source,
                "facet_id": facet_id,
                "aq_id": aq_id,
            })

            if exists_row:
                updated += 1
            else:
                inserted += 1

        except Exception as e:
            skipped += 1
            errors.append(
                schemas.QSSWUploadRowError(
                    row=line_no,
                    error=str(e),
                    raw={k: (r.get(k) or "") for k in reader.fieldnames},
                )
            )

    if not dry_run:
        db.commit()

    return schemas.QSSWUploadResult(
        ok=True,
        dry_run=dry_run,
        inserted=inserted,
        updated=updated,
        skipped=skipped,
        errors=errors,
    )
# ============================================================
# 7. UPLOAD QUESTIONS (B1)
# ============================================================

@router.post(
    "/upload-questions",
    response_model=UploadQuestionsResult,
    summary="Bulk upload Questions via CSV (versioned)",
)
async def upload_questions(
    assessment_version: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    """
    B1 spec:
    - Admin-only (router dependency)
    - Input: multipart/form-data with assessment_version + CSV
    - Strict headers
    - Reads: skills
    - Writes: questions
    - Output: {uploaded:int, skipped:int, errors:list}
    - Non-idempotent: duplicates are skipped

    Notes:
    - `id` is INTERNAL DB PK (integer)
    - `question_code` is the CANONICAL / EXTERNAL identifier used by student flows
    - `prerequisite_qid` is INTEGER FK -> must be int/None
    """
    if not (file.filename or "").lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only .csv files are accepted")

    raw = await file.read()
    # Excel often writes UTF-8 with BOM -> utf-8-sig handles that safely.
    text = raw.decode("utf-8-sig", errors="replace")

    if not text.strip():
        raise HTTPException(status_code=400, detail="CSV file is empty")

    reader = csv.DictReader(io.StringIO(text))

    # Strict header validation: allow exactly one of the two supported sets (in exact order)
    expected_headers_a = [
        "id",
        "question_code",
        "question_text_en",
        "question_text_hi",
        "question_text_ta",
        "skill_id",
        "weight",
        "group_id",
        "prerequisite_qid",
    ]
    expected_headers_b = [
        "question_id",
        "question_code",
        "question_text_en",
        "question_text_hi",
        "question_text_ta",
        "skill_id",
        "weight",
        "group_id",
        "prerequisite_qid",
    ]

    actual_headers = reader.fieldnames or []
    if actual_headers != expected_headers_a and actual_headers != expected_headers_b:
        raise HTTPException(
            status_code=400,
            detail=f"Header mismatch. Expected exactly {expected_headers_a} OR {expected_headers_b} but got {actual_headers}",
        )

    uploaded = 0
    skipped = 0
    errors: List[str] = []

    for row_num, row in enumerate(reader, start=2):
        # Normalize id field (support both id and question_id)
        qid = (row.get("id") or row.get("question_id") or "").strip()
        if not qid:
            skipped += 1
            errors.append(f"Row {row_num}: missing id/question_id")
            continue

        # DB PK is integer
        try:
            qid_int = int(qid)
        except ValueError:
            skipped += 1
            errors.append(f"Row {row_num}: question id must be integer (got '{qid}')")
            continue

        # Non-idempotent behavior: skip if PK already exists
        if db.query(models.Question).filter_by(id=qid_int).first():
            skipped += 1
            errors.append(f"Row {row_num}: duplicate question id '{qid_int}'")
            continue

        # B2 shared validation (no DB write before validation passes)
        row_for_validation = dict(row)
        row_for_validation["assessment_version"] = (assessment_version or "").strip()

        validated = validate_question_row(db=db, row=row_for_validation, row_index=row_num)
        if isinstance(validated, RowValidationError):
            skipped += 1
            for fe in validated.errors:
                errors.append(f"Row {row_num}: {fe.field} - {fe.message}")
            continue

        # weight (optional default 1)
        weight_raw = (row.get("weight") or "1").strip()
        try:
            weight = int(weight_raw)
        except Exception:
            skipped += 1
            errors.append(f"Row {row_num}: weight must be an integer (got '{weight_raw}')")
            continue

        # prerequisite_qid is INTEGER FK -> must be int/None
        try:
            prereq_int = _parse_optional_int(row.get("prerequisite_qid"))
        except ValueError:
            skipped += 1
            errors.append(f"Row {row_num}: prerequisite_qid must be an integer (or empty/null)")
            continue

        q = models.Question(
            id=qid_int,
            assessment_version=validated.assessment_version,
            question_text_en=validated.question_text_en,
            question_text_hi=(validated.question_text_hi or "").strip() or None,
            question_text_ta=(validated.question_text_ta or "").strip() or None,
            skill_id=validated.skill_id,
            weight=weight,
            group_id=(row.get("group_id") or "").strip() or None,
            prerequisite_qid=prereq_int,
            question_code=(row.get("question_code") or "").strip() or None,
        )

        db.add(q)
        uploaded += 1

    db.commit()
    logger.info(f"upload-questions: uploaded={uploaded}, skipped={skipped}, errors={len(errors)}")
    return {"uploaded": uploaded, "skipped": skipped, "errors": errors}


# ============================================================
# PR1. UPLOAD ASSOCIATED QUALITIES (AQ_MASTER)
# ============================================================
@router.post(
    "/upload-aqs",
    response_model=UploadResponse,
    summary="Bulk upload Associated Qualities (AQ) via CSV",
)
async def upload_aqs(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    """
    PR1 (versioned) expected CSV headers (exact):
      assessment_version,aq_code,name_en,name_hi,name_ta,status

    Required:
      - assessment_version
      - aq_code
      - name_en

    Writes to:
      - associated_qualities_v (versioned knowledge pack table)

    Upsert behavior:
      - Unique key: (assessment_version, aq_code)
      - If exists => update name/status fields
      - Else insert
    """
    if file.content_type != "text/csv":
        raise HTTPException(status_code=400, detail="Only text/csv files are accepted")

    text_csv = (await file.read()).decode("utf-8-sig")
    if not text_csv.strip():
        raise HTTPException(status_code=400, detail="CSV file is empty")

    reader = csv.DictReader(io.StringIO(text_csv))

    expected_headers = ["assessment_version", "aq_code", "name_en", "name_hi", "name_ta", "status"]
    actual_headers = reader.fieldnames or []
    if actual_headers != expected_headers:
        raise HTTPException(
            status_code=400,
            detail=f"Header mismatch. Expected exactly {expected_headers} but got {actual_headers}",
        )

    inserted = 0
    updated = 0
    skipped = 0
    warnings: List[str] = []

    for row_num, row in enumerate(reader, start=2):
        assessment_version = (row.get("assessment_version") or "").strip()
        aq_code = (row.get("aq_code") or "").strip()
        original_aq_code = aq_code
        aq_code, applied = resolve_alias(
            db=db,
            entity_type="AQ",
            raw_value=aq_code,
            assessment_version=assessment_version,
        )
        if applied:
            warnings.append(
                f"Row {row_num}: aq_code '{original_aq_code}' normalized to '{aq_code}'"
            )
        name_en = (row.get("name_en") or "").strip()

        name_hi = (row.get("name_hi") or "").strip() or None
        name_ta = (row.get("name_ta") or "").strip() or None
        status = (row.get("status") or "").strip() or "active"

        if not assessment_version or not aq_code or not name_en:
            skipped += 1
            continue

        # Pre-check existence to keep inserted vs updated counts accurate
        exists = db.execute(
            text(
                """
                SELECT 1
                FROM associated_qualities_v
                WHERE assessment_version = :v AND aq_code = :c
                LIMIT 1
                """
            ),
            {"v": assessment_version, "c": aq_code},
        ).first()

        db.execute(
            text(
                """
                INSERT INTO associated_qualities_v
                    (assessment_version, aq_code, name_en, name_hi, name_ta, status)
                VALUES
                    (:assessment_version, :aq_code, :name_en, :name_hi, :name_ta, :status)
                ON CONFLICT (assessment_version, aq_code)
                DO UPDATE SET
                    name_en = EXCLUDED.name_en,
                    name_hi = EXCLUDED.name_hi,
                    name_ta = EXCLUDED.name_ta,
                    status  = EXCLUDED.status,
                    updated_at = NOW()
                """
            ),
            {
                "assessment_version": assessment_version,
                "aq_code": aq_code,
                "name_en": name_en,
                "name_hi": name_hi,
                "name_ta": name_ta,
                "status": status,
            },
        )

        if exists:
            updated += 1
        else:
            inserted += 1

    db.commit()
    logger.info(f"upload-aqs(v): inserted={inserted}, updated={updated}, skipped={skipped}")
    return {
        "status": "success",
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "warnings": warnings,
    }

# ============================================================
# PR1. UPLOAD AQ FACETS (AQ_FACET_TAXONOMY) - VERSIONED
# ============================================================

@router.post(
    "/upload-aq-facets",
    response_model=UploadResponse,
    summary="Bulk upload AQ Facets via CSV",
)
async def upload_aq_facets(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    """
    PR1 (versioned) expected CSV headers (exact):
      assessment_version,facet_code,aq_code,name_en,name_hi,name_ta,description_en,description_hi,description_ta,status

    Required:
      - assessment_version
      - facet_code
      - aq_code
      - name_en

    Writes to:
      - aq_facets_v (versioned knowledge pack table)

    FK behavior (version-safe):
      - (assessment_version, aq_code) must exist in associated_qualities_v
        otherwise row is skipped (and not inserted)

    Upsert behavior:
      - Unique key: (assessment_version, facet_code)
      - If exists => update fields
      - Else insert
    """
    if file.content_type != "text/csv":
        raise HTTPException(status_code=400, detail="Only text/csv files are accepted")

    raw = await file.read()

    # Robust decoding: prefer UTF-8 (BOM-safe), fallback to UTF-16 (Excel/Windows common)
    try:
        text_csv = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text_csv = raw.decode("utf-16")
        except UnicodeDecodeError:
            raise HTTPException(
                status_code=400,
                detail="Unable to decode CSV. Please save as CSV UTF-8 (Comma delimited).",
            )
    if not text_csv.strip():
        raise HTTPException(status_code=400, detail="CSV file is empty")

    reader = csv.DictReader(io.StringIO(text_csv))

    expected_headers = [
        "assessment_version",
        "facet_code",
        "aq_code",
        "name_en",
        "name_hi",
        "name_ta",
        "description_en",
        "description_hi",
        "description_ta",
        "status",
    ]
    actual_headers = reader.fieldnames or []
    if actual_headers != expected_headers:
        raise HTTPException(
            status_code=400,
            detail=f"Header mismatch. Expected exactly {expected_headers} but got {actual_headers}",
        )

    inserted = 0
    updated = 0
    skipped = 0
    warnings: List[str] = []

    for row_num, row in enumerate(reader, start=2):
        assessment_version = (row.get("assessment_version") or "").strip()
        facet_code = (row.get("facet_code") or "").strip()
        original_facet_code = facet_code  # <-- MOVE HERE (this is the exact fix)

        # PR42: normalize FACET codes like FACET01 -> FACET_01
        m = re.match(r"^(FACET)(\d{1,3})$", facet_code, flags=re.IGNORECASE)
        if m:
            facet_code = f"{m.group(1).upper()}_{int(m.group(2)):02d}"

        aq_code = (row.get("aq_code") or "").strip()

        facet_code, facet_applied = resolve_alias(
            db=db,
            entity_type="FACET",
            raw_value=facet_code,
            assessment_version=assessment_version,
        )
        if facet_applied:
            warnings.append(
                f"Row {row_num}: facet_code '{original_facet_code}' normalized to '{facet_code}'"
            )

        original_aq_code = aq_code
        aq_code, aq_applied = resolve_alias(
            db=db,
            entity_type="AQ",
            raw_value=aq_code,
            assessment_version=assessment_version,
        )
        if aq_applied:
            warnings.append(
                f"Row {row_num}: aq_code '{original_aq_code}' normalized to '{aq_code}'"
            )
        name_en = (row.get("name_en") or "").strip()

        name_hi = (row.get("name_hi") or "").strip() or None
        name_ta = (row.get("name_ta") or "").strip() or None

        description_en = (row.get("description_en") or "").strip() or None
        description_hi = (row.get("description_hi") or "").strip() or None
        description_ta = (row.get("description_ta") or "").strip() or None

        status = (row.get("status") or "").strip() or "active"

        if not assessment_version or not facet_code or not aq_code or not name_en:
            skipped += 1
            continue

        # FK check (version-safe): AQ must exist in associated_qualities_v for same version
        aq_exists = db.execute(
            text(
                """
                SELECT 1
                FROM associated_qualities_v
                WHERE assessment_version = :v AND aq_code = :aq
                LIMIT 1
                """
            ),
            {"v": assessment_version, "aq": aq_code},
        ).first()

        if not aq_exists:
            skipped += 1
            continue

        # Pre-check existence for accurate inserted vs updated counts
        exists = db.execute(
            text(
                """
                SELECT 1
                FROM aq_facets_v
                WHERE assessment_version = :v AND facet_code = :fc
                LIMIT 1
                """
            ),
            {"v": assessment_version, "fc": facet_code},
        ).first()

        db.execute(
            text(
                """
                INSERT INTO aq_facets_v
                    (assessment_version, facet_code, aq_code, name_en, name_hi, name_ta,
                     description_en, description_hi, description_ta, status)
                VALUES
                    (:assessment_version, :facet_code, :aq_code, :name_en, :name_hi, :name_ta,
                     :description_en, :description_hi, :description_ta, :status)
                ON CONFLICT (assessment_version, facet_code)
                DO UPDATE SET
                    aq_code = EXCLUDED.aq_code,
                    name_en = EXCLUDED.name_en,
                    name_hi = EXCLUDED.name_hi,
                    name_ta = EXCLUDED.name_ta,
                    description_en = EXCLUDED.description_en,
                    description_hi = EXCLUDED.description_hi,
                    description_ta = EXCLUDED.description_ta,
                    status = EXCLUDED.status,
                    updated_at = NOW()
                """
            ),
            {
                "assessment_version": assessment_version,
                "facet_code": facet_code,
                "aq_code": aq_code,
                "name_en": name_en,
                "name_hi": name_hi,
                "name_ta": name_ta,
                "description_en": description_en,
                "description_hi": description_hi,
                "description_ta": description_ta,
                "status": status,
            },
        )

        if exists:
            updated += 1
        else:
            inserted += 1

    db.commit()
    logger.info(f"upload-aq-facets(v): inserted={inserted}, updated={updated}, skipped={skipped}")
    return {
        "status": "success",
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "warnings": warnings,
    }


# ============================================================
# PR2. UPLOAD QUESTION ↔ AQ_FACET TAGGING (QUESTION_AQ_FACET_TAGGING)
# ============================================================

@router.post(
    "/upload-question-facet-tags",
    response_model=UploadResponse,
    summary="Bulk upload Question ↔ AQ Facet tags via CSV (versioned)",
)
async def upload_question_facet_tags(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UserSchema = Depends(get_current_active_user),
):
    """
    Expected CSV headers (exact):
      assessment_version,question_code,facet_code,tag_weight

    Behavior:
    - Upsert into question_facet_tags_v using UNIQUE(assessment_version, question_code, facet_code)
    - Validation:
      - question must exist in questions for that (assessment_version, question_code)
      - facet must exist in aq_facets_v for that (assessment_version, facet_code)
    """
    if file.content_type != "text/csv":
        raise HTTPException(status_code=400, detail="Only text/csv files are accepted")

    raw = await file.read()

    # Robust decoding: prefer UTF-8 (BOM-safe), fallback to UTF-16 (Excel/Windows common)
    try:
        text_csv = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text_csv = raw.decode("utf-16")
        except UnicodeDecodeError:
            raise HTTPException(
                status_code=400,
                detail="Unable to decode CSV. Please save as CSV UTF-8 (Comma delimited).",
            )
    if not text_csv.strip():
        raise HTTPException(status_code=400, detail="CSV file is empty")

    reader = csv.DictReader(io.StringIO(text_csv))

    expected_fields = {"assessment_version", "question_code", "facet_code", "tag_weight"}
    if set(reader.fieldnames or []) != expected_fields:
        raise HTTPException(
            status_code=400,
            detail=f"CSV must have columns exactly {expected_fields}, but got {reader.fieldnames}",
        )

    inserted = 0
    updated = 0
    skipped = 0
    warnings: List[str] = []

    for row_num, row in enumerate(reader, start=2):
        assessment_version = (row.get("assessment_version") or "").strip()
        question_code = (row.get("question_code") or "").strip()
        facet_code = (row.get("facet_code") or "").strip()
        original_facet_code = facet_code

        # PR42 — deterministic normalization for AQ facet codes:
        # Accepts variants like: AQ3F2, AQ03F2, AQ3_F2, aq3f2 -> AQ03_F2
        normalized_facet_code = facet_code.strip().upper().replace(" ", "")

        m = re.match(r"^AQ_?(\d{1,2})_?F_?(\d{1,2})$", normalized_facet_code)
        if m:
            aq_num = int(m.group(1))
            f_num = int(m.group(2))
            normalized_facet_code = f"AQ{aq_num:02d}_F{f_num}"
        # else: leave as-is (it will be validated against aq_facets_v)

        # PR42 — alias resolution second (optional, but consistent with PR42 framework)
        facet_code, facet_applied = resolve_alias(
            db=db,
            entity_type="FACET",
            raw_value=normalized_facet_code,
            assessment_version=assessment_version,
        )

        if normalized_facet_code != original_facet_code.upper().replace(" ", "") or facet_applied:
            warnings.append(
                f"Row {row_num}: facet_code '{original_facet_code}' → '{facet_code}'"
            )
        tag_weight_raw = (row.get("tag_weight") or "").strip()

        if not assessment_version or not question_code or not facet_code or not tag_weight_raw:
            skipped += 1
            continue

        try:
            tag_weight = int(tag_weight_raw)
        except ValueError:
            skipped += 1
            continue

        # Validate question exists (authoritative questions table)
        q = (
            db.query(Question)
            .filter_by(assessment_version=assessment_version, question_code=question_code)
            .first()
        )
        if not q:
            skipped += 1
            continue

        # Validate facet exists (versioned facets table from PR1)
        facet_exists = db.execute(
            text(
                """
                SELECT 1
                FROM aq_facets_v
                WHERE assessment_version = :v
                  AND facet_code = :f
                LIMIT 1
                """
            ),
            {"v": assessment_version, "f": facet_code},
        ).fetchone()

        if not facet_exists:
            skipped += 1
            continue

        # Upsert into versioned mapping table
        # We count "inserted" as successful upserts to match existing UploadResponse schema.
        result = db.execute(
            text(
                """
                INSERT INTO question_facet_tags_v
                    (assessment_version, question_code, facet_code, tag_weight)
                VALUES
                    (:v, :q, :f, :w)
                ON CONFLICT (assessment_version, question_code, facet_code)
                DO UPDATE SET
                    tag_weight = EXCLUDED.tag_weight,
                    updated_at = NOW()
                """
            ),
            {"v": assessment_version, "q": question_code, "f": facet_code, "w": tag_weight},
        )

        # NOTE: Postgres doesn't easily tell insert vs update here without RETURNING tricks.
        # We keep updated/skipped counters for logs; "inserted" means "upserted".
        inserted += 1

    db.commit()
    logger.info(
        f"upload-question-facet-tags(v): inserted={inserted}, updated={updated}, skipped={skipped}"
    )
    return {"status": "success", "inserted": inserted, "updated": updated, "skipped": skipped, "warnings": warnings}


# ============================================================
# PR3. UPLOAD AQ STUDENTSKILL WEIGHTS
# ============================================================

@router.post("/upload-aq-studentskill-weights")
async def upload_aq_studentskill_weights(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    PR3: Upload AQ -> StudentSkill weights (versioned bridge).

    Required CSV headers:
      assessment_version, aq_code, skill_id, weight
    Optional:
      status (defaults to 'active')
    """
    raw = await file.read()
    text_csv = raw.decode("utf-8-sig")

    reader = csv.DictReader(StringIO(text_csv))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV missing header row")

    headers = {h.strip() for h in reader.fieldnames}
    required = {"assessment_version", "aq_code", "skill_id", "weight"}
    missing = required - headers
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {sorted(missing)}. Found: {sorted(headers)}",
        )

    rows = []
    errors = []
    warnings: List[str] = []

    # Parse + basic validation
    for line_no, r in enumerate(reader, start=2):
        av = (r.get("assessment_version") or "").strip()
        aq = (r.get("aq_code") or "").strip()
        original_aq = aq
        aq, applied = resolve_alias(
            db=db,
            entity_type="AQ",
            raw_value=aq,
            assessment_version=av,
        )
        if applied:
            warnings.append(
                f"Row {line_no}: aq_code '{original_aq}' normalized to '{aq}'"
            )
        status = (r.get("status") or "active").strip() or "active"

        # skill_id
        try:
            skill_id = int((r.get("skill_id") or "").strip())
        except Exception:
            errors.append({"row": line_no, "field": "skill_id", "message": "skill_id must be an integer"})
            continue

        # weight
        try:
            weight = float((r.get("weight") or "").strip())
        except Exception:
            errors.append({"row": line_no, "field": "weight", "message": "weight must be numeric"})
            continue

        if not av:
            errors.append({"row": line_no, "field": "assessment_version", "message": "assessment_version is required"})
            continue
        if not aq:
            errors.append({"row": line_no, "field": "aq_code", "message": "aq_code is required"})
            continue
        if weight < 0:
            errors.append({"row": line_no, "field": "weight", "message": "weight must be >= 0"})
            continue

        rows.append(
            {
                "rownum": line_no,
                "assessment_version": av,
                "aq_code": aq,
                "skill_id": skill_id,
                "weight": weight,
                "status": status,
            }
        )

    if errors:
        raise HTTPException(status_code=400, detail={"ok": False, "errors": errors})

    # Duplicate detection inside file
    seen = set()
    for r in rows:
        k = (r["assessment_version"], r["aq_code"], r["skill_id"])
        if k in seen:
            errors.append({"row": r["rownum"], "field": "duplicate", "message": "Duplicate key in file (assessment_version, aq_code, skill_id)"})
        seen.add(k)

    if errors:
        raise HTTPException(status_code=400, detail={"ok": False, "errors": errors})

    # Validate AQ exists (versioned)
    aq_keys = sorted({(r["assessment_version"], r["aq_code"]) for r in rows})
    # Build VALUES list safely via parameters
    # We’ll query by version and aq_code separately to avoid tuple-IN quirks.
    versions = sorted({v for v, _ in aq_keys})
    aq_codes = sorted({c for _, c in aq_keys})

    aq_ok_rows = db.execute(
        text("""
            SELECT assessment_version, aq_code
            FROM associated_qualities_v
            WHERE status='active'
              AND assessment_version = ANY(:versions)
              AND aq_code = ANY(:aq_codes)
        """),
        {"versions": versions, "aq_codes": aq_codes},
    ).fetchall()
    aq_ok = set((x[0], x[1]) for x in aq_ok_rows)

    for r in rows:
        if (r["assessment_version"], r["aq_code"]) not in aq_ok:
            errors.append({"row": r["rownum"], "field": "aq_code", "message": "AQ not found for assessment_version in associated_qualities_v (active)"})

    # Validate skill exists
    skill_ids = sorted({r["skill_id"] for r in rows})
    skill_ok_rows = db.execute(
        text("SELECT id FROM skills WHERE id = ANY(:ids)"),
        {"ids": skill_ids},
    ).fetchall()
    skill_ok = {x[0] for x in skill_ok_rows}

    for r in rows:
        if r["skill_id"] not in skill_ok:
            errors.append({"row": r["rownum"], "field": "skill_id", "message": "skill_id not found in skills"})

    if errors:
        raise HTTPException(status_code=400, detail={"ok": False, "errors": errors})

    # Reject if any already exist in DB (beta-safety)
    # We’ll check by version + aq_code + skill_id using a join on a temp VALUES set
    keys = [(r["assessment_version"], r["aq_code"], r["skill_id"]) for r in rows]

    # Build a VALUES table using SQLAlchemy parameters
    # This pattern avoids huge OR clauses and stays deterministic.
    values_sql = ", ".join([f"(:v{i}, :a{i}, :s{i})" for i in range(len(keys))])
    params = {}
    for i, (v, a, s) in enumerate(keys):
        params[f"v{i}"] = v
        params[f"a{i}"] = a
        params[f"s{i}"] = s

    existing = db.execute(
        text(f"""
            SELECT w.assessment_version, w.aq_code, w.skill_id
            FROM aq_student_skill_weights w
            JOIN (VALUES {values_sql}) AS x(assessment_version, aq_code, skill_id)
              ON w.assessment_version = x.assessment_version
             AND w.aq_code = x.aq_code
             AND w.skill_id = x.skill_id
        """),
        params,
    ).fetchall()

    if existing:
        # Return deterministic list
        already = [{"assessment_version": v, "aq_code": a, "skill_id": s} for (v, a, s) in existing]
        raise HTTPException(status_code=400, detail={"ok": False, "errors": [{"row": None, "field": "duplicate_db", "message": "Some keys already exist in DB", "existing": already}]})

    # Insert
    inserted = 0
    for r in rows:
        db.execute(
            text("""
                INSERT INTO aq_student_skill_weights (assessment_version, aq_code, skill_id, weight, status)
                VALUES (:assessment_version, :aq_code, :skill_id, :weight, :status)
            """),
            r,
        )
        inserted += 1

    db.commit()
    return {"ok": True, "inserted": inserted, "errors": [], "warnings": warnings}

# ============================================================
# PR16. upload_explainability_language_pack
# ============================================================

@router.post("/upload-explainability-language-pack", response_model=ExplainabilityUploadResult)
def upload_explainability_language_pack(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    PR16: Admin bulk upload for explainability CMS copy.
    - Versioned + locale-aware upsert by (version, locale, explanation_key)
    - Student-safe text only (no analytics)
    - CSV decoded with utf-8-sig to handle Excel BOM
    """
    # Basic file validation
    if not file.filename.lower().endswith(".csv"):
        return ExplainabilityUploadResult(
            total_rows=0, inserted=0, updated=0, skipped=0,
            errors=[ExplainabilityUploadRowError(row=0, error="Only .csv files are supported")]
        )

    raw = file.file.read()
    try:
        decoded = raw.decode("utf-8-sig")
    except Exception:
        return ExplainabilityUploadResult(
            total_rows=0, inserted=0, updated=0, skipped=0,
            errors=[ExplainabilityUploadRowError(row=0, error="Unable to decode file as utf-8 / utf-8-sig")]
        )

    reader = csv.DictReader(io.StringIO(decoded))

    required_cols = {"version", "locale", "explanation_key", "text"}
    if not reader.fieldnames or not required_cols.issubset(set([h.strip() for h in reader.fieldnames])):
        return ExplainabilityUploadResult(
            total_rows=0, inserted=0, updated=0, skipped=0,
            errors=[ExplainabilityUploadRowError(
                row=0,
                error=f"Missing required columns. Required: {sorted(list(required_cols))}"
            )]
        )

    inserted = 0
    updated = 0
    skipped = 0
    errors = []
    total_rows = 0

    # Helper: normalize boolean
    def parse_bool(val):
        if val is None:
            return True
        s = str(val).strip().lower()
        if s == "":
            return True
        if s in {"true", "1", "yes", "y"}:
            return True
        if s in {"false", "0", "no", "n"}:
            return False
        # unknown values -> treat as True but record warning via skipped? keep minimal
        return True

    # Process rows (row index = 2 because DictReader row1 is header)
    for idx, row in enumerate(reader, start=2):
        total_rows += 1

        version = (row.get("version") or "").strip()
        locale = (row.get("locale") or "").strip()
        explanation_key = (row.get("explanation_key") or "").strip()
        text_val = (row.get("text") or "").strip()
        is_active = parse_bool(row.get("is_active"))

        # Validate required fields
        if not version or not locale or not explanation_key or not text_val:
            skipped += 1
            errors.append(ExplainabilityUploadRowError(
                row=idx,
                error="Missing one of required fields: version, locale, explanation_key, text"
            ))
            continue

        # Upsert by unique key
        existing = db.execute(
            select(ExplainabilityContent).where(
                ExplainabilityContent.version == version,
                ExplainabilityContent.locale == locale,
                ExplainabilityContent.explanation_key == explanation_key,
            )
        ).scalars().first()

        if existing:
            existing.text = text_val
            existing.is_active = is_active
            updated += 1
        else:
            db.add(ExplainabilityContent(
                version=version,
                locale=locale,
                explanation_key=explanation_key,
                text=text_val,
                is_active=is_active,
            ))
            inserted += 1

    db.commit()

    return ExplainabilityUploadResult(
        total_rows=total_rows,
        inserted=inserted,
        updated=updated,
        skipped=skipped,
        errors=errors,
    )

@router.post(
    "/upload-skill-aliases",
    response_model=UploadResponse,
    summary="PR42: Upload alias → canonical mappings (AQ/FACET/SKILL) via CSV",
)
async def upload_skill_aliases(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    # Content-type gate (consistent with other endpoints)
    if file.content_type not in ("text/csv", "application/vnd.ms-excel"):
        raise HTTPException(status_code=400, detail="Only text/csv files are accepted")

    text_csv = (await file.read()).decode("utf-8-sig")
    if not text_csv.strip():
        raise HTTPException(status_code=400, detail="CSV file is empty")

    reader = csv.DictReader(io.StringIO(text_csv))

    expected_headers = [
        "entity_type",
        "assessment_version",
        "alias",
        "canonical_code",
        "is_active",
        "notes",
    ]
    actual_headers = reader.fieldnames or []
    if actual_headers != expected_headers:
        raise HTTPException(
            status_code=400,
            detail=f"Header mismatch. Expected exactly {expected_headers} but got {actual_headers}",
        )

    def parse_bool(v: str, default: bool = True) -> bool:
        if v is None:
            return default
        s = str(v).strip().lower()
        if s == "":
            return default
        if s in ("true", "t", "1", "yes", "y"):
            return True
        if s in ("false", "f", "0", "no", "n"):
            return False
        raise ValueError(f"Invalid boolean value '{v}' (expected true/false)")

    inserted = 0
    updated = 0
    skipped = 0
    warnings: List[str] = []
    errors = []

    allowed_types = {"AQ", "FACET", "SKILL"}

    for row_num, row in enumerate(reader, start=2):
        try:
            entity_type = (row.get("entity_type") or "").strip().upper()
            assessment_version = (row.get("assessment_version") or "").strip() or None
            alias = (row.get("alias") or "").strip()
            canonical_code = (row.get("canonical_code") or "").strip()
            is_active = parse_bool(row.get("is_active"), default=True)
            notes = (row.get("notes") or "").strip() or None

            if entity_type not in allowed_types:
                raise ValueError(f"entity_type must be one of {sorted(list(allowed_types))}")

            if not alias or not canonical_code:
                skipped += 1
                warnings.append(f"Row {row_num}: skipped (alias/canonical_code missing)")
                continue

            # Redundant alias (same as canonical after normalization) → skip with warning
            if norm(alias) == norm(canonical_code):
                skipped += 1
                warnings.append(
                    f"Row {row_num}: skipped (alias '{alias}' is same as canonical '{canonical_code}' after normalization)"
                )
                continue

            # Locate existing mapping (case-insensitive alias)
            existing = (
                db.query(models.SkillAlias)
                .filter(
                    models.SkillAlias.entity_type == entity_type,
                    models.SkillAlias.assessment_version.is_(None)
                    if assessment_version is None
                    else models.SkillAlias.assessment_version == assessment_version,
                    func.lower(models.SkillAlias.alias) == alias.lower(),
                )
                .first()
            )

            if existing:
                # Conflict protection: same alias cannot map to different canonical
                if norm(existing.canonical_code) != norm(canonical_code):
                    raise ValueError(
                        f"Conflict: alias '{alias}' already maps to '{existing.canonical_code}', not '{canonical_code}'"
                    )

                # Update only non-key fields (idempotent)
                changed = False
                if existing.is_active != is_active:
                    existing.is_active = is_active
                    changed = True
                if notes is not None and existing.notes != notes:
                    existing.notes = notes
                    changed = True

                # Keep canonical_code stable; allow casing/format cleanup if equivalent
                if existing.canonical_code != canonical_code and norm(existing.canonical_code) == norm(canonical_code):
                    existing.canonical_code = canonical_code
                    changed = True

                if changed:
                    updated += 1
                else:
                    skipped += 1

            else:
                db.add(
                    models.SkillAlias(
                        entity_type=entity_type,
                        assessment_version=assessment_version,
                        alias=alias,
                        canonical_code=canonical_code,
                        is_active=is_active,
                        notes=notes,
                    )
                )
                inserted += 1

        except Exception as e:
            errors.append({"row": row_num, "error": str(e), "raw": dict(row)})

    if errors:
        raise HTTPException(status_code=400, detail={"ok": False, "errors": errors})

    try:
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"DB constraint error: {str(e)}")

    return {
        "status": "success",
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "warnings": warnings,
    }

